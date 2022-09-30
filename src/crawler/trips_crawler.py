#!/usr/bin/python
# coding:utf-8
import sys
import os
import datetime
import openpyxl
import urllib.request
import urllib
import re
import json
import time
import threading
import shapefile as shp
import pandas as pd
import numpy as np
import random
import zipfile
from argparse import Namespace
from math import radians, cos, sin, asin, sqrt
import schedule
from coordTransform_py import CoordTransform_utils


class RoutePlanning(object):
    def __init__(self, name, inputWorkbook='trip_mobike_20190108.xlsx', keysBook='amap_keys.xlsx', inputCoordType='wgs84', outputCoordType='wgs84', planningType='walking', convertor=True, zipFiles=False, test=False, time_internal=15):
        # {"status":"0","info":"USER_DAILY_QUERY_OVER_LIMIT","infocode":"10044"}
        self.projectName = name
        self.zipFiles = zipFiles
        self.deleteProcessFile = True
        self.planningType = planningType
        self.inputCoordType = inputCoordType
        self.outputCoordType = outputCoordType
        self.send_headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/'
            '43.0.2357.81 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Connection': 'keep-alive',
            'Cookie': r'xxxxxx',
        }
        
        # input-related attributes
        self.keys = []
        keysSheet = openpyxl.load_workbook(keysBook).get_sheet_by_name("Sheet1")
        for row in range(2, keysSheet.max_row+1):
            self.keys.append(keysSheet.cell(row, 2).value)
        self.sheet = openpyxl.load_workbook(
            inputWorkbook).get_sheet_by_name("Sheet1")
        self.maxRow, self.cols = self.sheet.max_row, self.sheet.max_column
        
        if test:
            self.maxRow = 100
        self.time_internal = time_internal
        
        # tabel_header
        self.tabel_header = []
        for item in range(1, self.cols+1):
            self.tabel_header.append(self.sheet.cell(1, item).value)
        self.id_index = self.tabel_header.index("ID")
        self.lon_0_index, self.lat_0_index = self.tabel_header.index("lon_0")+1, self.tabel_header.index("lat_0")+1
        self.lon_1_index, self.lat_1_index = self.tabel_header.index("lon_1")+1, self.tabel_header.index("lat_1")+1

        # self.datetime = time.strftime('%y%m%d%H%M%S',time.localtime(time.time()))
        self.convertor = convertor
        self.time_crawl_data = time.strftime('%d%H%M', time.localtime(time.time()))
        self.storePath = 'D:/' if sys.platform == 'win32' else '/home/pcl/Data/GBA/new/'

        # storage-related attibutes
        self.save_points_whole_trip = False
        self.save_trip_detail = True
        self.save_trip_step_detail = True
        self.save_trip_tmcs_detail = False
        # shapefile strorage
        self.save_Trajectory = True
        self.save_trajectory_step = True
        self.save_trajectory_tmcs = True

        self.file_initial()

        pass

    def file_initial(self):
        self.datetime = time.strftime('%y%m%d', time.localtime(time.time()))
        if self.save_trip_detail:
            self.csv_trip_info = open(
                self.storePath+self.projectName+'_trajectory_info_'+self.datetime+'.csv', mode='w')
            self.csv_trip_info.write(
                "tripID,duration,travelDis,speed,smooth,slow,congested,verycongested,unknown,toll_distance,tolls,traffic_lights,strategy" + '\n')

        if self.save_points_whole_trip:
            self.csv_Trajectory_points = open(
                self.storePath+self.projectName+'_route_'+self.datetime+'.csv', mode='w')
            self.csv_Trajectory_points.write("tripID,nodeID,lon,lat")

        if self.save_Trajectory:
            self.shapefile_whole_trip = shp.Writer(
                self.storePath+self.projectName+'_路径规划_'+self.datetime)
            self.shapefile_whole_trip.field('tripID', 'C', size=20)
            self.shapefile_whole_trip.field('duration', 'N')
            self.shapefile_whole_trip.field('travelDis', 'N')
            self.shapefile_whole_trip.field('speed', 'N', decimal=1)
            self.shapefile_whole_trip.field('smooth', 'N', decimal=3)
            self.shapefile_whole_trip.field('slow', 'N', decimal=3)
            self.shapefile_whole_trip.field('congested', 'N', decimal=3)
            self.shapefile_whole_trip.field('verycongested', 'N', decimal=3)
            self.shapefile_whole_trip.field('unknown', 'N', decimal=3)
            self.shapefile_whole_trip.field('toll_distance', 'N')
            self.shapefile_whole_trip.field('tolls', 'N')
            self.shapefile_whole_trip.field('traffic_lights', 'N')
            self.shapefile_whole_trip.field('strategy', 'C', size=50)
            self.shapefile_whole_trip.autoBalance = 1

        if self.save_trip_tmcs_detail:
            # self.file_tmcs_tra = open(self.storePath+self.projectName+'tmcs_轨迹_'+self.datetime+'.csv', mode='w')
            self.csv_trip_tmcs_info = open(
                self.storePath+self.projectName+'_tmcs_'+self.datetime+'.csv', mode='w')
            # self.file_tmcs_tra.write("tripID,nodeID,lon,lat")
            self.csv_trip_tmcs_info.write("tripID,distance,status")

        if self.save_trajectory_tmcs:
            self.shapefile_tmcs = shp.Writer(
                self.storePath+self.projectName+'_tmcs_'+self.datetime)
            self.shapefile_tmcs.field('lineID', 'C', size=20)
            self.shapefile_tmcs.field('tripID', 'C', size=20)
            self.shapefile_tmcs.field('distance', 'N')
            self.shapefile_tmcs.field('status', 'N')

        if self.save_trajectory_step:
            self.shapefile_steps = shp.Writer(
                self.storePath+self.projectName+'_step_'+self.datetime)
            self.shapefile_steps.field('ID', 'C', size=20)
            self.shapefile_steps.field('tripID','C', size=20)
            # self.shapefile_steps.field('instruction','c',size=250)
            self.shapefile_steps.field('road', 'C', size=30)
            self.shapefile_steps.field('distance', 'N')
            self.shapefile_steps.field('duration', 'N')
            self.shapefile_steps.field('tolls', 'N')

        if self.save_trip_step_detail:
            self.csv_trip_step_detail = open(
                self.storePath+self.projectName+'_step_'+self.datetime+'.csv', mode='w')
            self.csv_trip_step_detail.write(
                "tripID,instruction,orientation,road,distance,tolls,toll_distance,toll_road,duration,action,assistant_action" + '\n')

        pass

    def json_to_object(self, data):
        # 将任意一个json字符串,转换为python的object对象,转换后支持使用属性访问, http://blog.csdn.net/chuanqi305/article/details/54846825
        return json.loads(data.decode("utf-8"), object_hook=lambda d: Namespace(**d), encoding='utf-8')

    # Get OD info from excel file
    def OD_Info(self, rowIndex):
        x_0, y_0 = round(self.sheet.cell(row=rowIndex, column=self.lon_0_index).value, 7), round(
            float(self.sheet.cell(row=rowIndex, column=self.lat_0_index).value), 6)
        x_1, y_1 = round(self.sheet.cell(row=rowIndex, column=self.lon_1_index).value, 7), round(
            float(self.sheet.cell(row=rowIndex, column=self.lat_1_index).value), 6)
        
        if self.inputCoordType == 'wgs84':
            [x_0, y_0] = CoordTransform_utils.wgs84_to_gcj02(x_0, y_0)
            [x_1, y_1] = CoordTransform_utils.wgs84_to_gcj02(x_1, y_1)
        elif self.inputCoordType == 'bd09':
            [x_0, y_0] = CoordTransform_utils.bd09_to_gcj02(x_0, y_0)
            [x_1, y_1] = CoordTransform_utils.bd09_to_gcj02(x_1, y_1)
        origin = str(round(x_0, 6))+","+str(round(y_0, 6))
        destination = str(round(x_1, 6))+","+str(round(y_1, 6))
        # return "&origin=" + origin + "&destination=" + destination
        return [origin, destination]

    # geocoding coordinate into adress
    def geocode(self, coordinate, key):
        geocode_url = "http://restapi.amap.com/v3/geocode/regeo?output=json&location=" + \
            coordinate+"&key=" + key + "&radius=500&extensions=base"
        req1 = urllib.request.Request(geocode_url, headers=self.send_headers)
        html1 = urllib.request.urlopen(req1).read()
        json_geo_data = self.json_to_object(data=html1)
        # json_geo_data = json.loads(html1, encoding='utf-8')
        return str.replace(json_geo_data.regeocode.formatted_address, '广东省深圳市', '')

    # act_o, act_d 已根据输入情况装换成相关的坐标体系
    def spideInfo(self, url, tripID, act_o, act_d):
        req = urllib.request.Request(url, headers=self.send_headers)
        try:
            html_raw = urllib.request.urlopen(req, timeout=30).read()
            # print(url, json_data)
            self.storeInfo(json=html_raw, tripID=tripID, act_o=act_o, act_d=act_d)
        except:
            pass

    # store data to txtfile
    def storeInfo(self, tripID, act_o, act_d, json):
        json_data = self.json_to_object(data=json)
        if self.planningType == 'walking':
            routeInfo = json_data.route
        elif self.planningType == 'bicycling':
            routeInfo = json_data.data
        elif self.planningType == 'driving':
            routeInfo = json_data.route

        if (self.save_Trajectory or self.save_trip_detail):
            index_of_node = 0
            x_pre, y_pre = 0.0, 0.0
            points_trajectory = []
            s1 = np.zeros(0)
            s2 = np.zeros(0)
            s3 = np.zeros(0)
            s4 = np.zeros(0)
            s5 = np.zeros(0)
            for item in routeInfo.paths[0].steps:
                for node in str.split(item.polyline, ';'):
                    [x, y] = [float(node.split(",")[0]),
                              float(node.split(",")[1])]
                    if self.outputCoordType == 'wgs84':
                        [x, y] = CoordTransform_utils.gcj02_to_wgs84(x, y)
                    elif self.outputCoordType == 'bd09':
                        [x, y] = CoordTransform_utils.gcj02_to_bd09(x, y)
                    if ((x != x_pre) & (y != y_pre)):
                        points_trajectory.append([x, y])
                        record = '\n%s,%d,%.6f,%.6f,' % (
                            (str(int(self.time_crawl_data)*1000 + tripID)), index_of_node, x, y)
                        index_of_node += 1
                        x_pre = x
                        y_pre = y
                        if self.save_points_whole_trip:
                            self.csv_Trajectory_points.write(record)

                for j, segment in enumerate(item.tmcs):
                    if segment.status == '畅通':
                        s1 = np.append(s1, float(segment.distance))
                    elif segment.status == '缓行':
                        s2 = np.append(s2, float(segment.distance))
                    elif segment.status == '拥堵':
                        s3 = np.append(s3, float(segment.distance))
                    elif segment.status == '非常拥堵':
                        s4 = np.append(s4, float(segment.distance))
                    else:
                        s5 = np.append(s5, float(segment.distance))
            dis = float(routeInfo.paths[0].distance)
            [s1_r, s2_r, s3_r, s4_r, s5_r] = [
                s1.sum()/dis, s2.sum()/dis, s3.sum()/dis, s4.sum()/dis, s5.sum()/dis]
            speed = str(dis/float(routeInfo.paths[0].duration))

            if self.save_trip_detail:
                # save info to csv file
                info = "%s,%s,%s,%s,%.3f,%.3f,%.3f,%.3f,%.3f,%s,%s,%s,%s" % ((str(int(self.time_crawl_data)*1000 + tripID)), routeInfo.paths[0].duration, dis, speed, (s1_r), (
                    s2_r), (s3_r), (s4_r), (s5_r), routeInfo.paths[0].toll_distance, routeInfo.paths[0].tolls, routeInfo.paths[0].traffic_lights, routeInfo.paths[0].strategy)
                self.csv_trip_info.write(info+"\n")
                print('\t'+info)
                pass

            # save the whole trajectory to shapefile
            if self.save_Trajectory:
                self.shapefile_whole_trip.line([points_trajectory])
                self.shapefile_whole_trip.record(str(tripID), str(routeInfo.paths[0].duration), str(
                    dis), speed, s1_r, s2_r, s3_r, s4_r, s5_r/dis, (routeInfo.paths[0].toll_distance), (routeInfo.paths[0].tolls), (routeInfo.paths[0].traffic_lights), routeInfo.paths[0].strategy)
                print("\tsave to shapefile")
                # self.shapefile_whole_trip.record( str((self.time_crawl_data)*1000 + tripID), routeInfo.paths[0].duration, str(dis), str(s1_r/dis), str(s2_r), str(s3_r), str(s4_r), str(s5_r),(routeInfo.paths[0].toll_distance), (routeInfo.paths[0].tolls), (routeInfo.paths[0].traffic_lights), routeInfo.paths[0].strategy )

        if (self.save_trajectory_step or self.save_trip_step_detail):
            for i, item in enumerate(routeInfo.paths[0].steps):
                if self.save_trip_step_detail:
                    temp = ('%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s' % (str(tripID)+"_"+str(i), item.instruction, item.orientation, item.road if hasattr(
                        item, 'road') else ' ', item.distance, item.tolls, item.toll_distance, item.toll_road, item.duration, item.action, item.assistant_action))
                    self.csv_trip_step_detail.write(temp+'\n')

                if self.save_trajectory_step:
                    points_trajectory = []
                    for node in str.split(item.polyline, ';'):
                        [x, y] = [float(node.split(",")[0]),
                                  float(node.split(",")[1])]
                        if self.outputCoordType == 'wgs84':
                            [x, y] = CoordTransform_utils.gcj02_to_wgs84(x, y)
                        elif self.outputCoordType == 'bd09':
                            [x, y] = CoordTransform_utils.gcj02_to_bd09(x, y)
                        points_trajectory.append([x, y])
                    self.shapefile_steps.line([points_trajectory])
                    self.shapefile_steps.record((self.time_crawl_data+str(tripID*100 + i)), tripID, (
                        item.road if hasattr(item, 'road') else ''), item.distance, item.duration, item.tolls)

        if (self.save_trajectory_tmcs or self.save_trip_tmcs_detail):
            for i, item in enumerate(routeInfo.paths[0].steps):
                for j, segment in enumerate(item.tmcs):
                    record_seg = "%s,%s,%s,%s,%s" % (str(tripID)+"_"+str(i)+"_"+str(j), time.time(), segment.distance, segment.status, segment.lcode)
                    if self.save_trip_tmcs_detail:
                        self.csv_trip_tmcs_info.write("\n"+record_seg)
                    if self.save_trajectory_tmcs:
                        points_trajectory = []
                        for node in str.split(segment.polyline, ';'):
                            [x, y] = [float(node.split(",")[0]),
                                      float(node.split(",")[1])]
                            if self.outputCoordType == 'wgs84':
                                [x, y] = CoordTransform_utils.gcj02_to_wgs84(
                                    x, y)
                            elif self.outputCoordType == 'bd09':
                                [x, y] = CoordTransform_utils.gcj02_to_bd09(
                                    x, y)
                            points_trajectory.append([x, y])
                        if segment.status == '畅通':
                            status = 1
                        elif segment.status == '缓行':
                            status = 2
                        elif segment.status == '拥堵':
                            status = 3
                        elif segment.status == '非常拥堵':
                            status = 4
                        else:
                            status = 0
                        self.shapefile_tmcs.line([points_trajectory])
                        self.shapefile_tmcs.record(
                            self.time_crawl_data+str(tripID*10000 + i*100 + j), tripID, segment.distance, status)

            pass

        pass

    def haversine(self, origin, destination):  # 经度1，纬度1，经度2，纬度2 （十进制度数）
        # Calculate the great circle distance between two points on the earth (specified in decimal degrees)
        lon1, lat1, lon2, lat2 = map(radians, [float(re.split(',|&', origin)[0]), float(re.split(
            ',|&', origin)[1]), float(re.split(',|&', destination)[0]), float(re.split(',|&', destination)[1])])
        # haversine公式
        dlon = lon2 - lon1
        dlat = lat2 - lat1
        a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
        c = 2 * asin(sqrt(a))
        r = 6371  # 地球平均半径，单位为公里
        return round(c * r * 1000, 2)

    def zipFile(self):
        files = []
        for parent,dirnames,filenames in os.walk(self.storePath):
            for filename in filenames:
                if (self.projectName in filename) & ('csv' not in filename):
                    files.append( self.storePath + filename )
        print(files)
        outputPath = 'D:' if sys.platform=='win32' else '/home/httpServer/'
        zip_name = outputPath + self.projectName +"_"+self.datetime+ '.zip'
        zip = zipfile.ZipFile(zip_name,'w', zipfile.ZIP_DEFLATED)
        try:
            for file in files:
                print ('compressing', file)
                zip.write(file)
            zip.close()
            print('compressing finished')
        except:
            print('zip files failed')

        if self.deleteProcessFile:
            try:
                for file in files:
                    os.remove(file)
            except:
                print('delete file failed')
        print(zip_name)
        return zip_name

    def start(self):
        # https://lbs.amap.com/api/webservice/guide/api/direction/#t8
        self.time_crawl_data = time.strftime(
            '%d%H%M', time.localtime(time.time()))
        # if(int(self.time_crawl_data[2:]) <= 3):
        if(int(self.time_crawl_data[2:]) >= 1730):
            self.shapefile_whole_trip.close()
            self.csv_trip_info.close()
            self.csv_trip_step_detail.close()
            self.file_initial()
            print("initial sucess!!!!!!!!!!!!!!!!!!!!!!!," +
                  str(self.time_crawl_data[3:]))
        else:
            print("initial nest time !!!!!!!!!!!!!!!!!!!!" +
                  str(self.time_crawl_data[2:]))
        threads_num = 1
        threads = []
        for i in range(2, self.maxRow + 1, threads_num):
            for j in range(threads_num):
                try:
                    row_index = i + j
                    tripID = self.sheet.cell(row=row_index, column=1).value
                    key = self.keys[tripID % len(self.keys)]
                    [origin, destination] = self.OD_Info(rowIndex=row_index)
                    if self.planningType == 'walking':
                        urlRoute = "https://restapi.amap.com/v3/direction/walking?key=" + \
                            key + "&origin=" + origin + "&destination=" + destination
                    elif self.planningType == 'bicycling':
                        urlRoute = "https://restapi.amap.com/v4/direction/bicycling?key=" + \
                            key + "&origin=" + origin + "&destination=" + destination
                    elif self.planningType == 'driving':
                        urlRoute = "https://restapi.amap.com/v3/direction/driving?key=" + \
                            key + "&origin=" + origin + "&destination=" + destination
                    print(urlRoute)
                    t = threading.Thread(target=self.spideInfo,  args=(
                        urlRoute, tripID, origin.replace(',', '&'), destination.replace(',', '&')))
                    threads.append(t)
                except:
                    # print('threads error')
                    pass
            for j in range(len(threads)):
                time.sleep(random.uniform(0, 1) / 8)
                threads[j].start()
            for j in range(len(threads)):
                threads[j].join()

            threads.clear()
        # if self.zipFiles:
        #     self.zipFile()
        # if self.save_trip_step_detail:
        #     self.csv_trip_step_detail.close()
        # if self.save_whole_trip_info:
        #     self.csv_trip_info.close()
        # if self.save_Trajectory:
        #     self.csv_Trajectory_points.close()

        if self.convertor:
            shapeFile = cf.CsvTransferToSHP()
            file_trajectory, file_details = self.projectName+'_轨迹_' + \
                self.datetime+'.csv', self.projectName+'_概况_'+self.datetime+'.csv'
            output = shapeFile.trans_polyline(
                self.storePath,  file_trajectory, file_details)
            # output_tmcs = shapeFile.trans_polyline(self.storePath,  file_trajectory_tmcs, file_details_tmcs )
            print('已压缩至文件：%s' % (output))
            return output

        # if self.save_trajectory_tmcs:
        #     self.shapefile_tmcs.close()
        # if self.save_Trajectory:
        #     self.shapefile_whole_trip.close()
        # if self.save_trajectory_step:
        #     self.shapefile_steps.close()
        #     print('close')
        # self.zipFile()
        
        pass

    def crawl_tasks(self):
        for hh in range(0, 24):
            if(hh < 6):
                for mm in range(0, 60, int(self.time_internal*2)):
                    schedule_time = f"{hh:02d}:{mm:02d}"
                    # print(schedule_time)
                    schedule.every().day.at(schedule_time).do(self.start)
            else:
                for mm in range(0, 60, int(self.time_internal)):
                    schedule_time = f"{hh:02d}:{mm:02d}"
                    # print(schedule_time)
                    schedule.every().day.at(schedule_time).do(self.start)

        while int(datetime.datetime.now().strftime("%H")) < 24:
            schedule.run_pending()
            time.sleep(30)
            print(datetime.datetime.now().strftime("%H:%M:%S"))
            pass


if __name__ == "__main__":
    routePlanning = RoutePlanning(name='GBA', inputWorkbook='../../db/cities_poi.xlsx', keysBook='../../db/amap_keys.xlsx', planningType='driving', inputCoordType='gcj02', convertor=False, test=False)
    # routePlanning = RoutePlanning(name='深圳市早高峰', inputWorkbook='d:/od_test.xlsx', keysBook='amap_keys.xlsx', planningType='driving', inputCoordType='wgs84', convertor=False,zipFiles=True, test=False)

    # routePlanning.start()
    routePlanning.crawl_tasks()

    # del RoutePlanning
    pass
